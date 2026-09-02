"""Local document scanning API powered by Docling and RapidOCR.

Run with: uvicorn app:app --reload --port 8000
Uploads are handled in a temporary local directory and are never sent to a cloud service.
"""
from __future__ import annotations

import re
import shutil
import tempfile
import threading
import zipfile
import os
from importlib.resources import files as package_files
from pathlib import Path

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

app = FastAPI(title="关键词检索器 API")
app.add_middleware(CORSMiddleware, allow_origins=["http://localhost:5173", "null", "file://"], allow_methods=["*"], allow_headers=["*"])
SUPPORTED = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".txt", ".md"}
MAX_ARCHIVE_FILES, MAX_ARCHIVE_BYTES = 300, 500 * 1024 * 1024
_text_converter = None
_ocr_converter = None
_scan_cancellations: dict[str, threading.Event] = {}
OFFLINE_MODEL_FILENAMES = {
    "det": "PP-OCRv6_det_small.onnx",
    "cls": "ch_ppocr_mobile_v2.0_cls_mobile.onnx",
    "rec": "PP-OCRv6_rec_small.onnx",
}

class ScanRequest(BaseModel):
    keywords: list[str] = Field(min_length=1)

class ScanResult(BaseModel):
    file: str
    page: int | None = None
    paragraph: int | None = None
    sheet: str | None = None
    row: int | None = None
    column: int | None = None
    keyword: str
    sentence: str

class CancelResponse(BaseModel):
    cancelled: bool

def safe_extract_zip(archive: Path, destination: Path) -> list[Path]:
    """Extract only supported regular files; reject zip-slip and archive bombs."""
    extracted: list[Path] = []
    with zipfile.ZipFile(archive) as bundle:
        infos = [info for info in bundle.infolist() if not info.is_dir()]
        if len(infos) > MAX_ARCHIVE_FILES or sum(info.file_size for info in infos) > MAX_ARCHIVE_BYTES:
            raise HTTPException(413, "压缩包包含过多或过大的文件")
        for info in infos:
            target = (destination / info.filename).resolve()
            if destination.resolve() not in target.parents or target.suffix.lower() not in SUPPORTED:
                continue
            target.parent.mkdir(parents=True, exist_ok=True)
            with bundle.open(info) as source, target.open("wb") as output:
                shutil.copyfileobj(source, output)
            extracted.append(target)
    return extracted

def sentence_hits(text: str, keyword: str) -> list[str]:
    return [sentence.strip() for sentence in re.split(r"(?<=[。！？.!?])\s*", text) if keyword.casefold() in sentence.casefold()]

def get_text_converter():
    """Create one reusable fast converter; never OCR normal Office/PDF text."""
    global _text_converter
    if _text_converter is None:
        from docling.document_converter import DocumentConverter, PdfFormatOption
        from docling.datamodel.base_models import InputFormat
        from docling.datamodel.pipeline_options import PdfPipelineOptions
        _text_converter = DocumentConverter(format_options={
            InputFormat.PDF: PdfFormatOption(pipeline_options=PdfPipelineOptions(do_ocr=False))
        })
    return _text_converter

def get_ocr_converter():
    """Create an OCR pipeline that only reads models shipped in this app."""
    global _ocr_converter
    if _ocr_converter is None:
        from docling.document_converter import DocumentConverter, PdfFormatOption
        from docling.datamodel.base_models import InputFormat
        from docling.datamodel.pipeline_options import PdfPipelineOptions, RapidOcrOptions
        asset_directory = os.environ.get("KEYWORD_SCANNER_ASSETS")
        model_root = Path(asset_directory) / "rapidocr-models" if asset_directory else Path(str(package_files("rapidocr").joinpath("models")))
        model_paths = {name: model_root / filename for name, filename in OFFLINE_MODEL_FILENAMES.items()}
        missing = [str(path) for path in model_paths.values() if not path.is_file()]
        if missing:
            raise HTTPException(503, f"离线 OCR 模型缺失：{', '.join(missing)}")
        options = PdfPipelineOptions(do_ocr=True, ocr_options=RapidOcrOptions(
            lang=["chinese"], det_model_path=str(model_paths["det"]), cls_model_path=str(model_paths["cls"]), rec_model_path=str(model_paths["rec"]),
        ))
        _ocr_converter = DocumentConverter(format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=options)})
    return _ocr_converter

def document_segments(document) -> list[tuple[str, int, int]]:
    """Keep every text item with its page and paragraph number within that page."""
    segments: list[tuple[str, int, int]] = []
    paragraphs_on_page: dict[int, int] = {}
    for item, _level in document.iterate_items():
        text = getattr(item, "text", None)
        if not isinstance(text, str) or not text.strip():
            continue
        provenance = getattr(item, "prov", [])
        page = provenance[0].page_no if provenance else 1
        paragraphs_on_page[page] = paragraphs_on_page.get(page, 0) + 1
        segments.append((text.strip(), page, paragraphs_on_page[page]))
    return segments

def parse_with_docling(source: Path) -> list[tuple[str, int, int]]:
    """Fast-path normal documents; OCR only image-only PDFs with no extracted text."""
    try:
        from docling.document_converter import DocumentConverter  # noqa: F401
    except ImportError as exc:
        raise HTTPException(503, "文档引擎未安装。请在 backend 目录安装 requirements.txt") from exc
    document = get_text_converter().convert(source).document
    segments = document_segments(document)
    if source.suffix.lower() == ".pdf" and len(" ".join(text for text, _, _ in segments).strip()) < 40:
        document = get_ocr_converter().convert(source).document
        segments = document_segments(document)
    if not segments:
        segments = [(document.export_to_markdown(), 1, 1)]
    return segments

def parse_excel(source: Path) -> list[tuple[str, str, int, int]]:
    """Return each populated spreadsheet cell with its worksheet, row, and column."""
    segments: list[tuple[str, str, int, int]] = []
    if source.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(source, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        segments.append((str(cell.value).strip(), worksheet.title, cell.row, cell.column))
        workbook.close()
        return segments
    import xlrd
    workbook = xlrd.open_workbook(source, on_demand=True)
    for worksheet in workbook.sheets():
        for row_index in range(worksheet.nrows):
            for column_index in range(worksheet.ncols):
                value = worksheet.cell_value(row_index, column_index)
                if str(value).strip():
                    segments.append((str(value).strip(), worksheet.name, row_index + 1, column_index + 1))
    workbook.release_resources()
    return segments

@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok", "engine": "Docling + RapidOCR"}

@app.post("/api/scan/{scan_id}/cancel", response_model=CancelResponse)
def cancel_scan(scan_id: str) -> CancelResponse:
    event = _scan_cancellations.get(scan_id)
    if event is None:
        return CancelResponse(cancelled=False)
    event.set()
    return CancelResponse(cancelled=True)

@app.post("/api/scan", response_model=list[ScanResult])
def scan(scan_id: str, keywords: str, files: list[UploadFile] = File(...)) -> list[ScanResult]:
    clean_keywords = [word.strip() for word in keywords.split(",") if word.strip()]
    if not clean_keywords:
        raise HTTPException(422, "请至少提供一个关键词")
    cancellation = threading.Event()
    _scan_cancellations[scan_id] = cancellation
    try:
        with tempfile.TemporaryDirectory(prefix="keyword-scanner-") as temporary:
            root, candidates = Path(temporary), []
            for upload in files:
                if cancellation.is_set():
                    raise HTTPException(409, "检测已中止")
                incoming = root / Path(upload.filename or "upload").name
                with incoming.open("wb") as output:
                    shutil.copyfileobj(upload.file, output)
                if incoming.suffix.lower() == ".zip":
                    candidates.extend(safe_extract_zip(incoming, root / f"expanded-{incoming.stem}"))
                elif incoming.suffix.lower() in SUPPORTED:
                    candidates.append(incoming)
            results: list[ScanResult] = []
            for source in candidates:
                if cancellation.is_set():
                    raise HTTPException(409, "检测已中止")
                if source.suffix.lower() in {".xls", ".xlsx"}:
                    for text, sheet, row, column in parse_excel(source):
                        if cancellation.is_set():
                            raise HTTPException(409, "检测已中止")
                        for keyword in clean_keywords:
                            for sentence in sentence_hits(text, keyword):
                                results.append(ScanResult(file=source.name, sheet=sheet, row=row, column=column, keyword=keyword, sentence=sentence))
                else:
                    for text, page, paragraph in parse_with_docling(source):
                        if cancellation.is_set():
                            raise HTTPException(409, "检测已中止")
                        for keyword in clean_keywords:
                            for sentence in sentence_hits(text, keyword):
                                results.append(ScanResult(file=source.name, page=page, paragraph=paragraph, keyword=keyword, sentence=sentence))
            return results
    finally:
        _scan_cancellations.pop(scan_id, None)
